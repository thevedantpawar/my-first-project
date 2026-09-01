#!/usr/bin/env python3
"""Combine tiers A, B and C into one working outbound sheet."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "output"
rows = []
for f, tier in [("send_tier_a.csv","A"), ("send_tier_b.csv","B"), ("review_tier_c.csv","C")]:
    rows += list(csv.DictReader(open(f"{OUT}/{f}", encoding="utf-8")))

# Tier A first, then B, then C; highest review volume first inside each tier
order = {"A":0, "B":1, "C":2}
rows.sort(key=lambda r: (order[r["tier"]], -int(r["review_count"] or 0)))

COLS = [
    ("Tier",          "tier",           7),
    ("Business",      "business_name",  38),
    ("City",          "city",           17),
    ("ST",            "state",          5),
    ("Email",         "email",          34),
    ("Phone",         "phone",          15),
    ("Website",       "website",        34),
    ("Rating",        "rating",         8),
    ("Reviews",       "review_count",   9),
    ("Closed days",   "closed_days",    18),
    ("Subject",       "subject",        17),
    ("First line (personalized)", "first_line", 62),
    ("Signal",        "signal_used",    22),
    ("Why this tier", "tier_reason",    38),
]
TRACK = [("Sent on", 11), ("Replied", 10), ("Notes", 30)]

ARIAL   = "Arial"
INK     = "1F2A28"
HEAD_BG = PatternFill("solid", fgColor="0F766E")
A_FILL  = PatternFill("solid", fgColor="E4F0E7")
B_FILL  = PatternFill("solid", fgColor="E2EFEC")
C_FILL  = PatternFill("solid", fgColor="F6EEDA")
FILL_BY = {"A":A_FILL, "B":B_FILL, "C":C_FILL}
INPUT_F = PatternFill("solid", fgColor="FFFF00")
THIN    = Side(style="thin", color="D5DAD8")
BORDER  = Border(bottom=THIN)

wb = Workbook()

# ---------------------------------------------------------------- Leads
ws = wb.active
ws.title = "Leads"

headers = [c[0] for c in COLS] + [t[0] for t in TRACK]
ws.append(headers)
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i)
    c.font = Font(name=ARIAL, bold=True, size=10, color="FFFFFF")
    c.fill = HEAD_BG
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
ws.row_dimensions[1].height = 30

for r in rows:
    line = []
    for _, key, _w in COLS:
        v = r[key]
        if key == "rating":
            v = float(v) if v else None
        elif key == "review_count":
            v = int(v) if v else 0
        line.append(v)
    ws.append(line + [None, None, None])

n = len(rows)
last = n + 1
for row in range(2, last + 1):
    tier = ws.cell(row=row, column=1).value
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name=ARIAL, size=10, color=INK)
        c.alignment = Alignment(vertical="top", wrap_text=(col == 12))
        c.border = BORDER
    ws.cell(row=row, column=1).fill = FILL_BY[tier]
    ws.cell(row=row, column=1).font = Font(name=ARIAL, size=10, bold=True, color=INK)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws.cell(row=row, column=8).number_format = "0.0"
    ws.cell(row=row, column=9).number_format = "#,##0"
    for col in (len(COLS)+1, len(COLS)+2, len(COLS)+3):
        ws.cell(row=row, column=col).fill = INPUT_F   # user fills these

for i, (_h, _k, w) in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for j, (_h, w) in enumerate(TRACK, len(COLS) + 1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"

# ---------------------------------------------------------------- Summary
sm = wb.create_sheet("Summary")
sm.column_dimensions["A"].width = 42
sm.column_dimensions["B"].width = 13
sm.column_dimensions["C"].width = 62

def put(r, a, b=None, c=None, bold=False, size=10):
    sm.cell(row=r, column=1, value=a).font = Font(name=ARIAL, bold=bold, size=size, color=INK)
    if b is not None:
        cell = sm.cell(row=r, column=2, value=b)
        cell.font = Font(name=ARIAL, bold=bold, size=size, color=INK)
        cell.number_format = "#,##0"
    if c is not None:
        cell = sm.cell(row=r, column=3, value=c)
        cell.font = Font(name=ARIAL, size=9, color="4A5B57")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

put(1, "Med spa outbound list", bold=True, size=14)
sm.cell(row=2, column=1,
        value="Tiers A, B and C combined. Suppressed rows excluded.").font = \
        Font(name=ARIAL, size=10, color="4A5B57")

put(4, "Tier", "Leads", "Meaning", bold=True)
put(5, "A  send first",   None, "Email domain matches the clinic's own website.")
put(6, "B  send second",  None, "Gmail, Yahoo and similar. Real addresses, often the owner personally.")
put(7, "C  verify first", None, "Corporate domain that does not match the website. Confirm it belongs to the clinic, not its web vendor.")
put(8, "Total in this file", None, None, bold=True)

counts = {t: sum(1 for r in rows if r["tier"] == t) for t in ("A", "B", "C")}
for r, tier in ((5, "A"), (6, "B"), (7, "C")):
    c = sm.cell(row=r, column=2, value=counts[tier])
    c.font = Font(name=ARIAL, size=10, color=INK); c.number_format = "#,##0"
c = sm.cell(row=8, column=2, value=sum(counts.values()))
c.font = Font(name=ARIAL, bold=True, size=10, color=INK); c.number_format = "#,##0"

put(10, "Not in this file", bold=True)
put(11, "Suppressed", 244, "202 scraper placeholders, 41 multi-location duplicates, 1 web agency address. Sending to these risks a roughly 19% bounce rate, which burns the sending domain.")

put(13, "Columns you fill in", bold=True)
put(14, "Sent on / Replied / Notes", None, "Shaded yellow on the Leads tab. Every other column is scraped or generated data.")

put(16, "How the first line was built", bold=True)
put(17, "Verified fields only", None, "Google rating, review count, published opening hours and service mix. The source file's has_online_booking column reads 'No' on 99.4% of rows and was wrong in every case checked against live sources, so it is used nowhere.")

put(19, "About these counts", bold=True)
put(20, "Counted from the rows in this file", None, "Written as values, not formulas. This is a fixed export of a completed audit, so the numbers do not need to recalculate.")

put(22, "Before sending", bold=True)
put(23, "Four blanks remain", None, "Sending address, agency URL, one real proof line, and a physical postal address for the CAN-SPAM footer.")

for r in range(1, 24):
    sm.row_dimensions[r].height = None
sm.row_dimensions[11].height = 44
sm.row_dimensions[17].height = 58
sm.row_dimensions[20].height = 32
sm.row_dimensions[23].height = 32
sm.row_dimensions[7].height = 32

path = f"{OUT}/medspa_leads_ABC.xlsx"
wb.save(path)
print("wrote", path, "with", n, "leads")
